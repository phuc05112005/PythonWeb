from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from datetime import datetime
from .models import (
    SANPHAM,
    LOAI,
    GIOHANG,
    CHITIETGIOHANG,
    DONHANG,
    CHITIETDONHANG,
    CUAHANG,
    TAIKHOAN,
    HINHANH,
    TONKHOSIZE,
    LICHSUKHO,
    SIZE_CHOICES,
)
from django.shortcuts import redirect
from .forms import UserRegisterForm, MailForm, CustomPasswordResetForm
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.db import transaction
from django.db.models.functions import TruncMonth
from functools import wraps
from django.contrib import messages

def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('dangnhap')
            
        # Tự động đồng bộ nếu chưa có bản ghi TAIKHOAN
        if not hasattr(request.user, 'taikhoan'):
            TAIKHOAN.objects.get_or_create(
                user=request.user,
                defaults={'role': 'admin' if request.user.is_staff else 'user'}
            )
            
        # Cho phép Admin và Quản lý truy cập các trang quản trị chung
        if request.user.taikhoan.role not in ['admin', 'quanly']:
            messages.error(request, "Bạn không có quyền truy cập trang này!")
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

from django.contrib.auth import authenticate, login, update_session_auth_hash
from django.contrib import messages
from django.core.paginator import Paginator
from .tool import cuahanggannhat
import os
import json
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime
import calendar
from django.utils.http import url_has_allowed_host_and_scheme

from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.conf import settings

SIZE_ORDER = [size for size, _ in SIZE_CHOICES]


def parse_non_negative_int(value, default=0):
    try:
        value_int = int(value)
    except (TypeError, ValueError):
        return default
    return max(value_int, 0)


def ensure_tonkho_records(sanpham, cuahang=None):
    """Đảm bảo mọi size đều có bản ghi tồn kho cho cửa hàng cụ thể hoặc tất cả cửa hàng"""
    if cuahang:
        cuahangs = [cuahang]
    else:
        cuahangs = CUAHANG.objects.all()
    
    for ch in cuahangs:
        tonkho_map = {row.size: row for row in sanpham.tonkho_sizes.filter(cuahang=ch)}
        for size in SIZE_ORDER:
            if size not in tonkho_map:
                TONKHOSIZE.objects.create(sanpham=sanpham, cuahang=ch, size=size, soluong=0)


def dong_bo_tong_ton_kho(sanpham):
    """Tổng tồn kho của sản phẩm = tổng tồn của tất cả cửa hàng"""
    tong = TONKHOSIZE.objects.filter(sanpham=sanpham).aggregate(total=Sum('soluong'))['total'] or 0
    if sanpham.soluong != tong:
        sanpham.soluong = tong
        sanpham.save(update_fields=['soluong'])
    return tong


def lay_ton_theo_size(sanpham, size, cuahang=None):
    """Lấy tồn của 1 size tại 1 cửa hàng. Nếu không có cuahang, trả về tổng toàn hệ thống."""
    query = TONKHOSIZE.objects.filter(sanpham=sanpham, size=size)
    if cuahang:
        tonkho = query.filter(cuahang=cuahang).first()
        if not tonkho:
            tonkho = TONKHOSIZE.objects.create(sanpham=sanpham, cuahang=cuahang, size=size, soluong=0)
        return tonkho.soluong
    else:
        return query.aggregate(total=Sum('soluong'))['total'] or 0


def doc_ton_kho_tu_post(post_data, cuahang_id=None):
    """Đọc dữ liệu tồn kho từ form. Nếu có cuahang_id, đọc ton_{ch_id}_{size}"""
    prefix = f"ton_{cuahang_id}_" if cuahang_id else "ton_"
    return {
        size: parse_non_negative_int(post_data.get(f'{prefix}{size.lower()}'), 0)
        for size in SIZE_ORDER
    }


def ghi_lich_su_kho(sanpham, cuahang, size, loai_biendong, so_luong, ton_truoc, ton_sau, user=None, ghichu=''):
    if so_luong <= 0:
        return
    LICHSUKHO.objects.create(
        sanpham=sanpham,
        cuahang=cuahang,
        size=size,
        loai_biendong=loai_biendong,
        soluong_thaydoi=so_luong,
        ton_truoc=ton_truoc,
        ton_sau=ton_sau,
        nguoithuchien=user,
        ghichu=ghichu,
    )


def home(request, loai_id=None):
    loai = LOAI.objects.all()
    if loai_id is None:
        sanpham = SANPHAM.objects.all()
    else:
        sanpham = SANPHAM.objects.filter(loaisp_id=loai_id)
    page = Paginator(sanpham, 9)
    page_number = request.GET.get('page')
    sanphamphantrang = page.get_page(page_number)
    context = {'sanpham': sanphamphantrang, 'loai': loai, 'loai_id': loai_id}
    return render(request, 'home.html', context)


def chitietsp(request, sanpham_id):
    sanpham = get_object_or_404(SANPHAM, id=sanpham_id)
    cuahangs = CUAHANG.objects.all()
    ensure_tonkho_records(sanpham)
    dong_bo_tong_ton_kho(sanpham)

    # Tồn kho theo từng cửa hàng
    tonkho_by_store = []
    for ch in cuahangs:
        inventory_items = []
        for s in SIZE_ORDER:
            item = sanpham.tonkho_sizes.filter(cuahang=ch, size=s).first()
            if not item:
                item = TONKHOSIZE.objects.create(sanpham=sanpham, cuahang=ch, size=s, soluong=0)
            inventory_items.append(item)
            
        total_stock = sum(inv.soluong for inv in inventory_items)
        tonkho_by_store.append({
            'store': ch,
            'inventory': inventory_items,
            'total_stock': total_stock
        })

    # Tồn kho tổng hợp theo size (để chọn size)
    tong_ton_size = sanpham.ton_theo_size
    tonkho_list = [(size, tong_ton_size.get(size, 0)) for size in SIZE_ORDER]

    context = {
        'sanpham': sanpham,
        'tonkho_by_store': tonkho_by_store,
        'tonkho_list': tonkho_list,
        'size_order': SIZE_ORDER,
        'cuahang': cuahangs,
    }
    return render(request, 'chitietsp.html', context)


def soluong(request):
    count = 0
    if request.user.is_authenticated:
        giohang, created = GIOHANG.objects.get_or_create(khachhang=request.user)
        if giohang:
            chitiet = CHITIETGIOHANG.objects.filter(giohang=giohang)
            for sp in chitiet:
                count += sp.soluong
    return {'count': count}


def dangky(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.is_active = False  # Chờ xác nhận email
            user.save()

            # Gửi email xác nhận
            current_site = get_current_site(request)
            subject = 'Kích hoạt tài khoản của bạn'
            message = render_to_string('activation_email.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)),
                'token': default_token_generator.make_token(user),
            })
            
            try:
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email])
                messages.success(request, "Vui lòng kiểm tra email để kích hoạt tài khoản!")
            except Exception as e:
                messages.error(request, f"Không thể gửi email kích hoạt: {e}")
            
            return redirect('dangnhap')
        else:
            messages.error(request, "Thông tin không hợp lệ!")
    else:
        form = UserRegisterForm()
    return render(request, 'dangky.html', {'form': form})


def kichhoat(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, "Tài khoản của bạn đã được kích hoạt thành công!")
        return redirect('dangnhap')
    else:
        messages.error(request, "Link kích hoạt không hợp lệ hoặc đã hết hạn!")
        return redirect('home')


def dangnhap(request):
    if request.user.is_authenticated:
        # Đồng bộ nhanh nếu chưa có TAIKHOAN
        if not hasattr(request.user, 'taikhoan'):
            TAIKHOAN.objects.get_or_create(
                user=request.user,
                defaults={'role': 'admin' if request.user.is_staff else 'user'}
            )
        if request.user.taikhoan.role == 'admin':
            return redirect('quantri')
        return redirect('home')

    if request.method == 'POST':
        taikhoan = request.POST.get('username')
        matkhau = request.POST.get('password')
        user = authenticate(request, username=taikhoan, password=matkhau)
        if user is not None:
            login(request, user)
            # Đồng bộ nhanh nếu chưa có TAIKHOAN
            if not hasattr(user, 'taikhoan'):
                TAIKHOAN.objects.get_or_create(
                    user=user,
                    defaults={'role': 'admin' if user.is_staff else 'user'}
                )
            if user.taikhoan.role == 'admin':
                return redirect('quantri')
            else:
                return redirect('home')
        else:
            messages.error(request, "Tên đăng nhập hoặc mật khẩu không chính xác!")
    return render(request, "dangnhap.html")


def thongtintaikhoan(request):
    if not request.user.is_authenticated:
        return redirect('dangnhap')

    if request.method == 'POST':
        # Xử lý cập nhật Email
        new_email = request.POST.get('email')
        if new_email:
            request.user.email = new_email
            request.user.save()
            messages.success(request, "Cập nhật email thành công.")

        # Xử lý đổi mật khẩu
        if 'old_password' in request.POST:
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Đổi mật khẩu thành công.")
                return redirect('thongtintaikhoan')
            else:
                messages.error(request, "Vui lòng kiểm tra lại thông tin đổi mật khẩu.")
        
        return redirect('thongtintaikhoan')
    else:
        password_form = PasswordChangeForm(request.user)

    context = {
        'password_form': password_form,
    }
    return render(request, 'taikhoan.html', context)


def themgiohang(request, sanpham_id):
    if not request.user.is_authenticated:
        return redirect('dangnhap')

    sanpham = get_object_or_404(SANPHAM, id=sanpham_id)
    size = (request.POST.get('size') or 'M').upper()
    soluongmua = parse_non_negative_int(request.POST.get('soluong'), 1)

    # Kiểm tra tổng tồn kho trước khi cho vào giỏ
    tong_ton = lay_ton_theo_size(sanpham, size)
    if soluongmua > tong_ton:
        messages.error(request, f"Sản phẩm hiện chỉ còn tổng {tong_ton} cái cho size {size}.")
        return redirect('chitiet', sanpham_id=sanpham_id)

    giohang, _ = GIOHANG.objects.get_or_create(khachhang=request.user)
    chitiet, created = CHITIETGIOHANG.objects.get_or_create(
        giohang=giohang,
        sanpham=sanpham,
        size=size,
        defaults={'soluong': soluongmua},
    )

    if not created:
        if (chitiet.soluong + soluongmua) > tong_ton:
            messages.error(request, f"Bạn đã có {chitiet.soluong} trong giỏ, không thể thêm {soluongmua} vì tổng kho chỉ có {tong_ton}.")
            return redirect('chitiet', sanpham_id=sanpham_id)
        chitiet.soluong += soluongmua
        chitiet.save()

    messages.success(request, "Đã thêm vào giỏ hàng.")
    return redirect('giohang')


def suagiohang(request, sanpham_id):
    if not request.user.is_authenticated:
        return redirect('dangnhap')

    if request.method == 'POST':
        size = (request.POST.get('size') or 'M').upper()
        soluongmoi = parse_non_negative_int(request.POST.get('soluong'), 1)
        gio_hang = GIOHANG.objects.get(khachhang=request.user)
        chitiet = CHITIETGIOHANG.objects.get(giohang=gio_hang, sanpham_id=sanpham_id, size=size)

        tong_ton = lay_ton_theo_size(chitiet.sanpham, size)
        if soluongmoi > tong_ton:
            messages.error(request, f"Tổng kho chỉ còn {tong_ton} cái.")
            return redirect('giohang')

        if soluongmoi <= 0:
            chitiet.delete()
        else:
            chitiet.soluong = soluongmoi
            chitiet.save()
    return redirect('giohang')


def xoagiohang(request, sanpham_id):
    if not request.user.is_authenticated:
        return redirect('dangnhap')

    gio_hang = GIOHANG.objects.get(khachhang=request.user)
    size = request.GET.get('size')

    CHITIETGIOHANG.objects.filter(giohang=gio_hang, sanpham_id=sanpham_id, size=size).delete()
    return redirect('giohang')


def giohang(request):
    if not request.user.is_authenticated:
        return redirect('dangnhap')
    gio_hang, created = GIOHANG.objects.get_or_create(khachhang=request.user)
    sanpham = CHITIETGIOHANG.objects.filter(giohang=gio_hang).select_related('sanpham')
    tongtien = 0
    for mon in sanpham:
        mon.tong_ton_size = lay_ton_theo_size(mon.sanpham, mon.size)
        tongtien += mon.sanpham.gia * mon.soluong
    context = {'sanpham': sanpham, 'tongtien': tongtien}
    return render(request, 'giohang.html', context)


def thanhtoan(request):
    if not request.user.is_authenticated:
        return redirect('dangnhap')
    
    gio_hang, created = GIOHANG.objects.get_or_create(khachhang=request.user)
    san_pham_gio = CHITIETGIOHANG.objects.filter(giohang=gio_hang).select_related('sanpham')
    
    if not san_pham_gio.exists():
        return redirect('giohang')

    tongtiendonhang = sum(mon.sanpham.gia * mon.soluong for mon in san_pham_gio)
    
    tienship = 0
    km = 0
    tong = tongtiendonhang

    if request.method == 'POST':
        lat_raw = request.POST.get('lat')
        lon_raw = request.POST.get('lon')
        action = request.POST.get('action')

        try:
            latkh = float(lat_raw) if lat_raw else None
            lonkh = float(lon_raw) if lon_raw else None
        except (TypeError, ValueError):
            latkh, lonkh = None, None

        if latkh is None or lonkh is None:
            messages.error(request, "Vui lòng xác định vị trí để tìm cửa hàng gần nhất.")
        else:
            cuahangs = CUAHANG.objects.all()
            if not cuahangs.exists():
                messages.error(request, "Hệ thống hiện không có cửa hàng nào hoạt động.")
                return redirect('giohang')

            # 1. Tìm danh sách các cửa hàng có ĐỦ HÀNG cho toàn bộ giỏ hàng
            cuahangs_hop_le = []
            for ch in CUAHANG.objects.all():
                du_hang = True
                for item in san_pham_gio:
                    tonkho = TONKHOSIZE.objects.filter(
                        sanpham=item.sanpham, size=item.size, cuahang=ch
                    ).first()
                    if not tonkho or tonkho.soluong < item.soluong:
                        du_hang = False
                        break
                if du_hang:
                    cuahangs_hop_le.append(ch)

            if not cuahangs_hop_le:
                messages.error(request, "Rất tiếc, hiện không có cửa hàng đơn lẻ nào có đủ tất cả sản phẩm trong giỏ hàng của bạn. Vui lòng liên hệ hotline để được hỗ trợ ghép đơn!")
                return redirect('giohang')

            # 2. Tìm cửa hàng GẦN NHẤT trong số các cửa hàng CÓ HÀNG
            ch_gan_nhat, km = cuahanggannhat(latkh, lonkh, cuahangs_hop_le)
            tienship = int(km * 15000)
            tong = tongtiendonhang + tienship

            if action == 'thanhtoan':
                with transaction.atomic():
                    # Khóa bản ghi để tránh tranh chấp (Race condition)
                    for item in san_pham_gio:
                        TONKHOSIZE.objects.select_for_update().get(
                            sanpham=item.sanpham, size=item.size, cuahang=ch_gan_nhat
                        )

                    # Tạo đơn hàng
                    don_hang = DONHANG.objects.create(
                        khachhang=request.user,
                        cuahang=ch_gan_nhat,
                        ten=request.POST.get('ten'),
                        sdt=request.POST.get('sdt'),
                        diachi=request.POST.get('diachi'),
                        lat=latkh,
                        lon=lonkh,
                        tongtien=tong,
                    )

                    for item in san_pham_gio:
                        CHITIETDONHANG.objects.create(
                            donhang=don_hang,
                            sanpham=item.sanpham,
                            size=item.size,
                            soluong=item.soluong,
                            dongia=item.sanpham.gia,
                        )
                        # Trừ kho cửa hàng cụ thể
                        tonkho = TONKHOSIZE.objects.get(sanpham=item.sanpham, size=item.size, cuahang=ch_gan_nhat)
                        ton_truoc = tonkho.soluong
                        tonkho.soluong -= item.soluong
                        tonkho.save()

                        ghi_lich_su_kho(
                            sanpham=item.sanpham,
                            cuahang=ch_gan_nhat,
                            size=item.size,
                            loai_biendong='xuat',
                            so_luong=item.soluong,
                            ton_truoc=ton_truoc,
                            ton_sau=tonkho.soluong,
                            user=request.user,
                            ghichu=f'Xuất kho cho đơn hàng #{don_hang.id}'
                        )
                        dong_bo_tong_ton_kho(item.sanpham)

                    CHITIETGIOHANG.objects.filter(giohang=gio_hang).delete()
                    return render(request, 'camon.html')

    context = {
        'sanpham': san_pham_gio,
        'tongtiendonhang': tongtiendonhang,
        'tongtien': tong,
        'km': km,
        'tienship': tienship,
        'tenkh': request.POST.get('ten') or f"{request.user.first_name} {request.user.last_name}",
        'sdtkh': request.POST.get('sdt'),
        'diachikh': request.POST.get('diachi'),
        'diachi_duong': request.POST.get('diachi_duong'),
        'diachi_phuong': request.POST.get('diachi_phuong'),
        'diachi_thanhpho': request.POST.get('diachi_thanhpho'),
        'latkh': request.POST.get('lat'),
        'lonkh': request.POST.get('lon'),
    }
    return render(request, 'thanhtoan.html', context)


def donhangcuatoi(request):
    if not request.user.is_authenticated:
        return redirect('dangnhap')
    donhang = DONHANG.objects.filter(khachhang=request.user).order_by('-ngaydat')
    return render(request, 'donhangcuatoi.html', {'donhang': donhang})


def chitietdonhang_khach(request, donhang_id):
    if not request.user.is_authenticated:
        return redirect('dangnhap')

    don_hang = get_object_or_404(DONHANG, id=donhang_id, khachhang=request.user)
    chitiet = CHITIETDONHANG.objects.filter(donhang=don_hang).select_related('sanpham')
    tong_san_pham = sum(item.dongia * item.soluong for item in chitiet)
    phi_giao_hang = max(don_hang.tongtien - tong_san_pham, 0)

    context = {
        'donhang': don_hang,
        'chitiet': chitiet,
        'tong_san_pham': tong_san_pham,
        'phi_giao_hang': phi_giao_hang,
    }
    return render(request, 'chitietdonhang_khach.html', context)


@admin_required
def quantri(request):
    tongsp = SANPHAM.objects.count()
    tongdh1 = DONHANG.objects.filter(trangthai=1).count()
    tongdh2 = DONHANG.objects.filter(trangthai=2).count()
    tongch = CUAHANG.objects.count()

    now = datetime.now()

    donhang = DONHANG.objects.filter(
        trangthai=2,
        ngaydat__month=now.month,
        ngaydat__year=now.year
    )

    tong = 0
    for dh in donhang:
        tong += dh.tongtien

    monthly_revenue = (
        DONHANG.objects.filter(trangthai=2)
        .annotate(month=TruncMonth('ngaydat'))
        .values('month')
        .annotate(total=Sum('tongtien'))
        .order_by('-month')
    )

    selected_month = request.GET.get('month')
    selected_month_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if selected_month:
        try:
            selected_month_date = datetime.strptime(selected_month, '%Y-%m')
        except ValueError:
            selected_month_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    max_day = calendar.monthrange(selected_month_date.year, selected_month_date.month)[1]
    day_from = request.GET.get('day_from')
    day_to = request.GET.get('day_to')
    day_from_value = ''
    day_to_value = ''

    try:
        if day_from is not None and day_from != '':
            day_from_value = max(1, min(int(day_from), max_day))
    except ValueError:
        day_from_value = ''

    try:
        if day_to is not None and day_to != '':
            day_to_value = max(1, min(int(day_to), max_day))
    except ValueError:
        day_to_value = ''

    monthly_details = DONHANG.objects.filter(
        trangthai=2,
        ngaydat__year=selected_month_date.year,
        ngaydat__month=selected_month_date.month
    ).order_by('-ngaydat')

    if day_from_value != '':
        monthly_details = monthly_details.filter(ngaydat__day__gte=day_from_value)
    if day_to_value != '':
        monthly_details = monthly_details.filter(ngaydat__day__lte=day_to_value)

    if day_from_value != '' and day_to_value != '' and day_from_value > day_to_value:
        day_from_value, day_to_value = day_to_value, day_from_value
        monthly_details = DONHANG.objects.filter(
            trangthai=2,
            ngaydat__year=selected_month_date.year,
            ngaydat__month=selected_month_date.month,
            ngaydat__day__gte=day_from_value,
            ngaydat__day__lte=day_to_value
        ).order_by('-ngaydat')

    selected_month_total = monthly_details.aggregate(total=Sum('tongtien'))['total'] or 0

    context = {
        'tongsp': tongsp,
        'tongdh1': tongdh1,
        'tongdh2': tongdh2,
        'tongcuahang': tongch,
        'tongtien': tong,
        'now': now,
        'monthly_revenue': monthly_revenue,
        'selected_month_date': selected_month_date,
        'monthly_details': monthly_details,
        'selected_month_total': selected_month_total,
        'day_from_value': day_from_value,
        'day_to_value': day_to_value,
        'max_day': max_day,
    }
    return render(request, 'quantri.html', context)


@admin_required
def quanlysanpham(request):
    sp = SANPHAM.objects.all().prefetch_related('tonkho_sizes').order_by('id')
    loai = LOAI.objects.all()

    keyword = request.GET.get('keyword')
    loai_id = request.GET.get('loai')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    tonkho = request.GET.get('tonkho')

    if keyword:
        sp = sp.filter(ten__icontains=keyword)
    if loai_id:
        sp = sp.filter(loaisp_id=loai_id)
    if min_price:
        sp = sp.filter(gia__gte=min_price)
    if max_price:
        sp = sp.filter(gia__lte=max_price)
    if tonkho == "con":
        sp = sp.filter(soluong__gt=0)
    elif tonkho == "het":
        sp = sp.filter(soluong=0)

    for item in sp:
        # Tồn kho theo size: tổng của tất cả cửa hàng
        tonkho_map = {size: 0 for size in SIZE_ORDER}
        for size in SIZE_ORDER:
            tonkho_map[size] = item.tonkho_sizes.filter(size=size).aggregate(total=Sum('soluong'))['total'] or 0
        
        item.tonkho_map = tonkho_map
        item.tong_ton = sum(tonkho_map.values())
        if item.soluong != item.tong_ton:
            SANPHAM.objects.filter(id=item.id).update(soluong=item.tong_ton)
            item.soluong = item.tong_ton

    context = {
        'sanpham': sp,
        'loai': loai,
        'size_order': SIZE_ORDER,
    }
    return render(request, 'admin/sanpham/index.html', context)


@admin_required
def xoasanpham(request, sanpham_id):
    if request.method == 'POST':
        sp = SANPHAM.objects.get(id=sanpham_id)
        if sp.hinh:
            if os.path.isfile(sp.hinh.path):
                os.remove(sp.hinh.path)
        sp.delete()
        messages.success(request, "Đã xóa thành công.")
    return redirect('quanlysanpham')


@admin_required
def suasanpham(request, sanpham_id):
    sp = SANPHAM.objects.get(id=sanpham_id)
    cuahangs = CUAHANG.objects.all()
    ensure_tonkho_records(sp)
    loai = LOAI.objects.all()

    if request.method == 'POST':
        sp.ten = request.POST.get('ten')
        sp.mota = request.POST.get('mota', '')
        loai_id = request.POST.get('loai')
        if loai_id:
            sp.loaisp_id = loai_id
        sp.gia = parse_non_negative_int(request.POST.get('gia'), sp.gia)

        hinhmoi = request.FILES.get('hinh')
        if hinhmoi:
            sp.hinh = hinhmoi

        # Cập nhật tồn kho cho TỪNG cửa hàng
        ghichu_input = request.POST.get('ghichu_chung', '').strip()
        final_ghichu = ghichu_input if ghichu_input else 'Cập nhật tồn kho khi sửa sản phẩm (đa cửa hàng)'

        for ch in cuahangs:
            ton_moi_theo_size = doc_ton_kho_tu_post(request.POST, cuahang_id=ch.id)
            for size in SIZE_ORDER:
                ton_row = TONKHOSIZE.objects.get(sanpham=sp, cuahang=ch, size=size)
                ton_cu = ton_row.soluong
                ton_moi = ton_moi_theo_size[size]
                if ton_moi == ton_cu:
                    continue
                
                loai_bd = 'dieuchinh_tang' if ton_moi > ton_cu else 'dieuchinh_giam'
                ghi_lich_su_kho(
                    sanpham=sp,
                    cuahang=ch,
                    size=size,
                    loai_biendong=loai_bd,
                    so_luong=abs(ton_moi - ton_cu),
                    ton_truoc=ton_cu,
                    ton_sau=ton_moi,
                    user=request.user,
                    ghichu=final_ghichu,
                )
                ton_row.soluong = ton_moi
                ton_row.save(update_fields=['soluong'])

        anhxoa = request.POST.getlist('anhxoa')
        if anhxoa:
            HINHANH.objects.filter(id__in=anhxoa).delete()

        hinhchitietmoi = request.FILES.getlist('hinhchitietmoi')
        for file in hinhchitietmoi:
            HINHANH.objects.create(sanpham=sp, hinh=file)

        sp.save()
        dong_bo_tong_ton_kho(sp)
        messages.success(request, "Cập nhật thành công.")
        return redirect('quanlysanpham')

    # Chuẩn bị dữ liệu cho template
    tonkho_data = []
    for ch in cuahangs:
        inventory_items = []
        for s in SIZE_ORDER:
            item = sp.tonkho_sizes.filter(cuahang=ch, size=s).first()
            if not item:
                item = TONKHOSIZE.objects.create(sanpham=sp, cuahang=ch, size=s, soluong=0)
            inventory_items.append(item)
            
        tonkho_data.append({
            'store': ch,
            'inventory_items': inventory_items
        })

    context = {
        'sanpham': sp,
        'dsloai': loai,
        'size_order': SIZE_ORDER,
        'tonkho_data': tonkho_data,
        'cuahangs': cuahangs,
    }
    return render(request, 'admin/sanpham/suasanpham.html', context)


@admin_required
def themsanpham(request):
    loaisp = LOAI.objects.all()
    cuahangs = CUAHANG.objects.all()
    
    if request.method == 'POST':
        tensp = request.POST.get('ten')
        motasp = request.POST.get('mota', '')
        loai = request.POST.get('loai')
        giasp = parse_non_negative_int(request.POST.get('gia'), 0)
        hinhsp = request.FILES.get('hinh')
        hinhchitietsp = request.FILES.getlist('hinhchitiet')

        sp = SANPHAM.objects.create(
            ten=tensp,
            mota=motasp,
            loaisp_id=loai,
            gia=giasp,
            soluong=0,
            hinh=hinhsp,
        )

        for ch in cuahangs:
            ton_theo_size = doc_ton_kho_tu_post(request.POST, cuahang_id=ch.id)
            for size in SIZE_ORDER:
                so_luong = ton_theo_size[size]
                TONKHOSIZE.objects.create(sanpham=sp, cuahang=ch, size=size, soluong=so_luong)
                if so_luong > 0:
                    ghi_lich_su_kho(
                        sanpham=sp,
                        cuahang=ch,
                        size=size,
                        loai_biendong='nhap',
                        so_luong=so_luong,
                        ton_truoc=0,
                        ton_sau=so_luong,
                        user=request.user,
                        ghichu='Nhập kho ban đầu khi tạo sản phẩm',
                    )

        dong_bo_tong_ton_kho(sp)

        for file in hinhchitietsp:
            HINHANH.objects.create(sanpham=sp, hinh=file)

        messages.success(request, "Lưu thành công.")
        return redirect('quanlysanpham')

    context = {'loai': loaisp, 'cuahangs': cuahangs, 'size_order': SIZE_ORDER}
    return render(request, 'admin/sanpham/themsanpham.html', context)


@admin_required
def lichsukho(request):
    lichsu = LICHSUKHO.objects.select_related('sanpham', 'nguoithuchien', 'cuahang')
    danhsach_sanpham = SANPHAM.objects.all().order_by('ten')

    sanpham_id = request.GET.get('sanpham')
    size = (request.GET.get('size') or '').upper()
    loai = request.GET.get('loai')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if sanpham_id:
        lichsu = lichsu.filter(sanpham_id=sanpham_id)
    if size in SIZE_ORDER:
        lichsu = lichsu.filter(size=size)
    if loai == 'tang':
        lichsu = lichsu.filter(loai_biendong__in=['nhap', 'dieuchinh_tang'])
    elif loai == 'giam':
        lichsu = lichsu.filter(loai_biendong__in=['xuat', 'dieuchinh_giam'])
    if from_date:
        lichsu = lichsu.filter(thoigian__date__gte=from_date)
    if to_date:
        lichsu = lichsu.filter(thoigian__date__lte=to_date)

    context = {
        'lichsu': lichsu,
        'danhsach_sanpham': danhsach_sanpham,
        'size_order': SIZE_ORDER,
    }
    return render(request, 'admin/kho/lichsu.html', context)


@admin_required
def quanlydonhang(request):
    donhang = DONHANG.objects.all()

    keyword = request.GET.get('keyword')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    status = request.GET.get('status')

    if keyword:
        donhang = donhang.filter(
            Q(ten__icontains=keyword) |
            Q(sdt__icontains=keyword)
        )
    if from_date:
        donhang = donhang.filter(ngaydat__date__gte=from_date)
    if to_date:
        donhang = donhang.filter(ngaydat__date__lte=to_date)
    if min_price:
        donhang = donhang.filter(tongtien__gte=min_price)
    if max_price:
        donhang = donhang.filter(tongtien__lte=max_price)
    if status == "0":
        donhang = donhang.filter(trangthai=1)  # chưa duyệt
    elif status == "1":
        donhang = donhang.filter(trangthai=2)  # đã duyệt

    donhang1 = donhang.filter(trangthai=1)
    donhang2 = donhang.filter(trangthai=2)

    return render(request, 'admin/donhang/index.html', {
        'donhang1': donhang1,
        'donhang2': donhang2,
    })


@admin_required
def duyetdonhang(request, donhang_id):
    if request.method == 'POST':
        dh = DONHANG.objects.get(id=donhang_id)
        dh.trangthai = 2
        dh.save()
        messages.success(request, "Cập nhật thành công.")
    return redirect('quanlydonhang')


@admin_required
def chitietdonhang(request, donhang_id):
    don_hang = DONHANG.objects.get(id=donhang_id)
    chitiet = CHITIETDONHANG.objects.filter(donhang=don_hang).select_related('sanpham')
    tong_san_pham = sum(item.dongia * item.soluong for item in chitiet)
    phi_giao_hang = max(don_hang.tongtien - tong_san_pham, 0)

    # Tính km từ tọa độ đã lưu
    km = None
    cuahang = CUAHANG.objects.all()
    if don_hang.lat and don_hang.lon and cuahang.exists():
        _, km_float = cuahanggannhat(don_hang.lat, don_hang.lon, cuahang)
        km = round(km_float, 2)

    context = {
        'chitiet': chitiet,
        'donhang': don_hang,
        'tong_san_pham': tong_san_pham,
        'phi_giao_hang': phi_giao_hang,
        'km': km,
    }
    return render(request, 'admin/donhang/chitietdonhang.html', context)


@admin_required
def xoadonhang(request, donhang_id):
    if request.method == 'POST':
        donhang = DONHANG.objects.get(id=donhang_id)
        donhang.delete()
        messages.success(request, "Xóa thành công.")
    return redirect('quanlydonhang')


@admin_required
def quanlyloai(request):
    loai = LOAI.objects.all()
    context = {'loai': loai}
    return render(request, 'admin/loaisp/index.html', context)


@admin_required
def themloai(request):
    if request.method == 'POST':
        loaisp = request.POST.get('loai')
        loai_sp = LOAI.objects.create(loai=loaisp)
        loai_sp.save()
        messages.success(request, "Thêm thành công.")
        return redirect('quanlyloai')
    return render(request, 'admin/loaisp/themloai.html')


@admin_required
def xoaloai(request, loai_id):
    if request.method == 'POST':
        loaisp = LOAI.objects.get(id=loai_id)
        loaisp.delete()
        messages.success(request, "Xóa thành công.")
    return redirect('quanlyloai')


@admin_required
def sualoai(request, loai_id):
    loaisp = LOAI.objects.get(id=loai_id)
    if request.method == 'POST':
        loaisp.loai = request.POST.get('loai')
        loaisp.save()
        messages.success(request, "Cập nhật thành công.")
        return redirect('quanlyloai')
    context = {'loai': loaisp}
    return render(request, 'admin/loaisp/sualoai.html', context)


def timkiem(request):
    loai = LOAI.objects.all()
    kq = SANPHAM.objects.none()
    search = ''
    if request.method == 'POST':
        search = request.POST.get('search', '')
        kq = SANPHAM.objects.filter(ten__unaccent__icontains=search)
    page = Paginator(kq, 9)
    page_number = request.GET.get('page')
    sanphamphantrang = page.get_page(page_number)
    context = {'kq': sanphamphantrang, 'search': search, 'loai': loai}
    return render(request, 'search.html', context)


@admin_required
def quanlycuahang(request):
    cuahang = CUAHANG.objects.all()
    context = {'cuahang': cuahang}
    return render(request, 'admin/cuahang/index.html', context)


@admin_required
def themcuahang(request):
    if request.method == 'POST':
        tench = request.POST.get('ten')
        sodienthoai = request.POST.get('sodienthoai', '').strip()
        diachich = request.POST.get('diachi')
        latch = request.POST.get('lat')
        lonch = request.POST.get('lon')
        if not latch or not lonch:
            messages.error(request, "Địa chỉ không hợp lệ.")
            context = {'tench': tench, 'sodienthoai': sodienthoai, 'diachich': diachich}
            return render(request, 'admin/cuahang/themcuahang.html', context)
        cuahang = CUAHANG.objects.create(
            ten=tench,
            sodienthoai=sodienthoai,
            diachi=diachich,
            lat=latch,
            lon=lonch,
            hinh=request.FILES.get('hinh'),
            gio_mo=request.POST.get('gio_mo') or "08:00",
            gio_dong=request.POST.get('gio_dong') or "22:00",
        )
        cuahang.save()
        messages.success(request, "Thêm thành công.")
        return redirect('quanlycuahang')
    return render(request, 'admin/cuahang/themcuahang.html')


@admin_required
def suacuahang(request, cuahang_id):
    cuahang = CUAHANG.objects.get(id=cuahang_id)
    if request.method == 'POST':
        cuahang.ten = request.POST.get('ten')
        cuahang.sodienthoai = request.POST.get('sodienthoai', '').strip()
        cuahang.diachi = request.POST.get('diachi')

        lat_str = request.POST.get('lat', cuahang.lat)
        lon_str = request.POST.get('lon', cuahang.lon)
        try:
            cuahang.lat = float(str(lat_str).replace(',', '.'))
            cuahang.lon = float(str(lon_str).replace(',', '.'))
        except ValueError:
            pass

        hinhmoi = request.FILES.get('hinh')
        if hinhmoi:
            cuahang.hinh = hinhmoi

        cuahang.gio_mo = request.POST.get('gio_mo') or cuahang.gio_mo
        cuahang.gio_dong = request.POST.get('gio_dong') or cuahang.gio_dong
        cuahang.save()
        messages.success(request, "Cập nhật cửa hàng thành công.")
        return redirect('quanlycuahang')

    context = {'cuahang': cuahang}
    return render(request, 'admin/cuahang/suacuahang.html', context)


@admin_required
def xoacuahang(request, cuahang_id):
    cuahang = CUAHANG.objects.get(id=cuahang_id)
    if request.method == 'POST':
        cuahang.delete()
        messages.success(request, "Xóa thành công.")
    return redirect('quanlycuahang')


def map_view(request):
    cuahang = CUAHANG.objects.all()
    data = []
    now = datetime.now().time()

    for ch in cuahang:
        is_open = ch.gio_mo <= now <= ch.gio_dong
        data.append({
            "ten": ch.ten,
            "sodienthoai": ch.sodienthoai,
            "lat": ch.lat,
            "lon": ch.lon,
            "diachi": ch.diachi,
            "hinh": ch.hinh.url if ch.hinh else "/static/images/cuahang1.jpg",
            "gio_mo": ch.gio_mo.strftime("%H:%M"),
            "gio_dong": ch.gio_dong.strftime("%H:%M"),
            "is_open": is_open,
        })
    return render(request, "map.html", {
        "stores_json": json.dumps(data, cls=DjangoJSONEncoder)
    })


@admin_required
def quanlytaikhoan(request):
    # Chỉ Admin mới được vào trang quản lý tài khoản
    if request.user.taikhoan.role != 'admin':
        messages.error(request, "Chỉ Admin mới có quyền quản lý tài khoản!")
        return redirect('quantri')

    # Tự động tạo TAIKHOAN cho các user cũ chưa có
    for u in User.objects.all():
        TAIKHOAN.objects.get_or_create(
            user=u,
            defaults={'role': 'admin' if u.is_staff else 'user'}
        )

    taikhoan = TAIKHOAN.objects.select_related('user').all().order_by('user__id')

    keyword = request.GET.get('keyword')
    role = request.GET.get('role')

    if keyword:
        taikhoan = taikhoan.filter(user__username__icontains=keyword)
    if role:
        taikhoan = taikhoan.filter(role=role)

    tong_taikhoan = TAIKHOAN.objects.count()
    tong_admin = TAIKHOAN.objects.filter(role='admin').count()
    tong_quanly = TAIKHOAN.objects.filter(role='quanly').count()
    tong_user = TAIKHOAN.objects.filter(role='user').count()

    context = {
        'taikhoan': taikhoan,
        'tong_taikhoan': tong_taikhoan,
        'tong_admin': tong_admin,
        'tong_quanly': tong_quanly,
        'tong_user': tong_user,
    }
    return render(request, 'admin/taikhoan/index.html', context)


@admin_required
def capnhatquyen(request, user_id):
    # Chỉ Admin mới có quyền cập nhật người khác
    if request.user.taikhoan.role != 'admin':
        messages.error(request, "Chỉ Admin mới có quyền thay đổi phân quyền!")
        return redirect('quantri')

    if request.method == 'POST':
        try:
            tk = TAIKHOAN.objects.select_related('user').get(user__id=user_id)
            # Không cho phép thay đổi superuser hoặc chính mình
            if tk.user.is_superuser:
                messages.error(request, "Không thể thay đổi quyền của Super Admin!")
                return redirect('quanlytaikhoan')
            if tk.user.id == request.user.id:
                messages.error(request, "Không thể thay đổi quyền của chính bạn!")
                return redirect('quanlytaikhoan')

            quyen_moi = request.POST.get('quyen')
            if quyen_moi in ['admin', 'quanly', 'user']:
                tk.role = quyen_moi
                # Đồng bộ is_staff
                tk.user.is_staff = quyen_moi in ['admin', 'quanly']
                tk.user.save()
                tk.save()
                messages.success(request, f'Đã cập nhật quyền cho {tk.user.username} thành {tk.get_role_display()}!')
        except TAIKHOAN.DoesNotExist:
            messages.error(request, "Tài khoản không tồn tại!")
    return redirect('quanlytaikhoan')


def xoataikhoan(request, id):
    if request.method == "POST":
        user = get_object_or_404(User, id=id)

        if user == request.user:
            messages.error(request, "Không thể xóa tài khoản của chính bạn!")
            return redirect('quanlytaikhoan')

        if user.is_superuser:
            messages.error(request, "Không thể xóa Super Admin!")
            return redirect('quanlytaikhoan')

        user.delete()
        messages.success(request, "Xóa tài khoản thành công!")

    return redirect('quanlytaikhoan')


from django.shortcuts import render, redirect
from django.core.mail import send_mail
from django.contrib import messages
from .forms import MailForm
from django.conf import settings


def send_mailtrap(request):
    if request.method == 'POST':
        form = MailForm(request.POST)
        if form.is_valid():
            email = request.POST.get('email')
            subject = request.POST.get('subject')
            message = request.POST.get('message')
            try:
                full_message = f"""
                Email khách: {email}

                Nội dung:
                {message}
                """

                send_mail(
                    subject,
                    full_message,
                    settings.DEFAULT_FROM_EMAIL,
                    ['phuc052005@gmail.com'],  # mail bạn nhận
                    fail_silently=False
                )
                messages.success(request, 'Email đã gửi thành công!')
            except Exception as e:
                messages.error(request, f'Gửi email thất bại: {e}')
            return redirect('send_mailtrap')
    else:
        form = MailForm()
    return render(request, 'send_mailtrap.html', {'form': form})


def gioithieu(request):
    return render(request, 'gioithieu.html')


def chinhsachbaohanh(request):
    return render(request, 'chinhsachbaohanh.html')


def chinhsachdoitra(request):
    return render(request, 'chinhsachdoitra.html')


def dieukhoansudung(request):
    return render(request, 'dieukhoansudung.html')

@admin_required
def export_inventory_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mau Cap Nhat Kho"

    # Header
    headers = ["ID SP", "Ten San Pham", "ID Cua Hang", "Ten Cua Hang", "Size", "Ton Hien Tai", "So Luong Them", "Ghi Chu"]
    ws.append(headers)

    # Data
    cuahangs = CUAHANG.objects.all()
    sanphams = SANPHAM.objects.all().order_by('id')

    for sp in sanphams:
        ensure_tonkho_records(sp)
        for ch in cuahangs:
            for s in SIZE_ORDER:
                ton = TONKHOSIZE.objects.filter(sanpham=sp, cuahang=ch, size=s).first()
                ws.append([sp.id, sp.ten, ch.id, ch.ten, s, ton.soluong if ton else 0, 0, ""])

    # Format
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mau_cap_nhat_kho.xlsx"'
    wb.save(response)
    return response


@admin_required
def import_inventory_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            success_count = 0
            # Duyệt từ dòng 2 (bỏ qua header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                sp_id, _, ch_id, _, size, _, so_luong_them, ghichu = row
                
                if so_luong_them and int(so_luong_them) != 0:
                    try:
                        sp = SANPHAM.objects.get(id=sp_id)
                        ch = CUAHANG.objects.get(id=ch_id)
                        ton_row, _ = TONKHOSIZE.objects.get_or_create(sanpham=sp, cuahang=ch, size=size)
                        
                        ton_cu = ton_row.soluong
                        so_luong_them = int(so_luong_them)
                        ton_moi = ton_cu + so_luong_them
                        
                        if ton_moi < 0: ton_moi = 0 # Không để âm
                        
                        # Ghi lịch sử
                        loai_bd = 'dieuchinh_tang' if so_luong_them > 0 else 'dieuchinh_giam'
                        ghi_lich_su_kho(
                            sanpham=sp,
                            cuahang=ch,
                            size=size,
                            loai_biendong=loai_bd,
                            so_luong=abs(so_luong_them),
                            ton_truoc=ton_cu,
                            ton_sau=ton_moi,
                            user=request.user,
                            ghichu=ghichu if ghichu else 'Cập nhật nhanh bằng Excel'
                        )
                        
                        ton_row.soluong = ton_moi
                        ton_row.save()
                        dong_bo_tong_ton_kho(sp)
                        success_count += 1
                    except Exception:
                        continue
            
            messages.success(request, f"Đã cập nhật thành công {success_count} bản ghi tồn kho.")
        except Exception as e:
            messages.error(request, f"Lỗi xử lý file Excel: {e}")
            
    return redirect('quanlysanpham')


def export_lichsukho_excel(request):
    lichsu = LICHSUKHO.objects.select_related('sanpham', 'nguoithuchien', 'cuahang')

    # --- GIỐNG FILTER TRANG HTML ---
    sanpham_id = request.GET.get('sanpham')
    size = (request.GET.get('size') or '').upper()
    loai = request.GET.get('loai')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if sanpham_id:
        lichsu = lichsu.filter(sanpham_id=sanpham_id)
    if size in SIZE_ORDER:
        lichsu = lichsu.filter(size=size)
    if loai == 'tang':
        lichsu = lichsu.filter(loai_biendong__in=['nhap', 'dieuchinh_tang'])
    elif loai == 'giam':
        lichsu = lichsu.filter(loai_biendong__in=['xuat', 'dieuchinh_giam'])
    if from_date:
        lichsu = lichsu.filter(thoigian__date__gte=from_date)
    if to_date:
        lichsu = lichsu.filter(thoigian__date__lte=to_date)

    # --- TẠO FILE EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = "LichSuKho"

    # Header
    ws.append([
        "Thời gian",
        "Sản phẩm",
        "Cửa hàng",
        "Size",
        "Loại",
        "Số lượng",
        "Tồn trước",
        "Tồn sau",
        "Người thao tác",
        "Ghi chú"
    ])

    # Data
    for item in lichsu:
        ws.append([
            item.thoigian.strftime("%d/%m/%Y %H:%M:%S"),
            item.sanpham.ten,
            item.cuahang.ten if item.cuahang else "-",
            item.size,
            item.get_loai_biendong_display(),
            item.soluong_thaydoi,
            item.ton_truoc,
            item.ton_sau,
            item.nguoithuchien.username if item.nguoithuchien else "-",
            item.ghichu or "-"
        ])

    # Response download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"lich_su_kho_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response